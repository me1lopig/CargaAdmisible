
# Funciones de obtención de datos
# 

# datos_terreno->Importa los datos incluidos en el archivo datos_terreno.xlsx
# datos_cimentacion->Importa los datos incluidos en el archivo datos_cimentacion.xlsx


# Importación de librerías
import numpy as np
import openpyxl



def datos_terreno():

    libro = openpyxl.load_workbook('datos_terreno.xlsx')
    hoja = libro.active
    
    # geometría de las capas
    espesor=[]
    cotas=[]
    # valores físicos
    pe_aparente=[]
    pe_saturada=[]
    # valores elásticos
    E=[]
    poisson=[]
    # valores de resistencia
    cohesion=[]
    fi=[]

    # tipos de datos
    tipo_datos=[]
   
    for row in hoja.iter_rows():
        espesor.append(row[0].value)
        espesor[0]=0
        az=sum(espesor) # vector de espesores
        nivel_freatico=hoja.cell(row=2, column=2).value
        pe_aparente.append(row[2].value)
        pe_aparente[0]=0
        pe_saturada.append(row[3].value)
        pe_saturada[0]=0
        E.append(row[4].value)
        E[0]=0
        poisson.append(row[5].value)
        poisson[0]=0
        cohesion.append(row[6].value)
        cohesion[0]=0
        fi.append(row[7].value)
        fi[0]=0


    for i in np.arange(len(espesor)):
        cotas.append(sum(espesor[0:i+1]))
    
    # se toman los datos de la cabecera
    for filas in hoja.iter_cols():
        tipo_datos.append(filas[0].value)

    return espesor,cotas,az,nivel_freatico,pe_aparente,pe_saturada,E,poisson,cohesion,fi




def datos_cimentacion():

    libro = openpyxl.load_workbook('datos_cimentacion.xlsx')
    hoja = libro.active
    
    # valores de definición de las caracteristicas de la cimentación
    b=hoja.cell(row=2, column=1).value
    l=hoja.cell(row=2, column=2).value
    forma=hoja.cell(row=2, column=3).value
    empotramiento=hoja.cell(row=2, column=4).value
    pendiente=hoja.cell(row=2, column=5).value
    axil=hoja.cell(row=2, column=6).value
    hb=hoja.cell(row=2, column=7).value
    hl=hoja.cell(row=2, column=8).value
    fs=hoja.cell(row=2, column=9).value
    nCalculos=hoja.cell(row=2, column=10).value
    incremento=hoja.cell(row=2, column=11).value


    return b,l,forma,empotramiento,pendiente,axil,hb,hl,fs,nCalculos,incremento