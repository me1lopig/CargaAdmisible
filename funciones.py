# Funciones auxiliares

# crea_directorio->Genera un directorio en el que guardar los cálculos
# datos_terreno->Importa los datos incluidos en el archivo datos_terreno.xlsx
# factoresCapacidad->factores de capacidad de carga
# correcciónForma->Coeficientes de corrección por la forma de la cimentación
# correccionTalud-> Coeficientes de corrección por cercanía de la cimentación a un talud
# densidad-> Cálculo de la densidad del terreno bajo la base de la cimentación


# Importación de librerías
import numpy as np
import openpyxl
import os
from datetime import datetime


def crea_directorio():
    # creación del directorio de trabajo para guardar resultados
    now = datetime.now()
    directorio=(str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second))
    os.mkdir(directorio)
    return directorio


def datos_terreno():

    libro = openpyxl.load_workbook('datos_terreno.xlsx')
    hoja = libro.active
    
    # geometría de las capas
    espesor=[]
    cotas=[]
    # valores físicos
    pe_seco=[]
    pe_saturado=[]
    # valores elásticos
    E=[]
    poisson=[]
    tipo_datos=[]
    # valores de resistencia
    cohesion=[]
    fi=[]
   
    for row in hoja.iter_rows():
        espesor.append(row[0].value)
        espesor[0]=0
        az=sum(espesor) # vector de espesores
        nivel_freatico=hoja.cell(row=2, column=2).value
        pe_seco.append(row[2].value)
        pe_seco[0]=0
        pe_saturado.append(row[3].value)
        pe_saturado[0]=0
        E.append(row[4].value)
        E[0]=0
        poisson.append(row[5].value)
        poisson[0]=0
        cohesion.append(row[6].value)
        cohesion[0]=0
        fi.append(row[7].value)
        fi[0]=0



    for i in np.arange(len(espesor)):
        cotas.append(sum(espesor[0:i+1]))
    
    # se toman los datos de la cabecera
    for filas in hoja.iter_cols():
        tipo_datos.append(filas[0].value)

    return espesor,cotas,az,nivel_freatico,pe_seco,pe_saturado,E,poisson,cohesion,fi



def factoresCapacidad (cohesion,anguloRozamiento):
     
    anguloRozamientoRad=np.deg2rad(anguloRozamiento)

    if (cohesion==0):
        Nc=(np.pi+2)
        Nq=1
        Ng=0
    else:
        Nq=((1+np.sin(anguloRozamientoRad))/(1-np.sin(anguloRozamientoRad)))*np.exp(np.pi*np.tan(anguloRozamientoRad))
        Nc=(Nq-1)/np.tan(anguloRozamientoRad)
        Ng=1.5*(Nq-1)*np.tan(anguloRozamientoRad)
    
    return Nc,Nq,Ng


def correccionForma(ancho,largo,anguloRozamiento):

    anguloRozamientoRad=np.deg2rad(anguloRozamiento)
    # factores por influencia de la forma
    sc=1+0.2*ancho/largo
    sq=1+1.5*np.tan(anguloRozamientoRad)*ancho/largo
    sg=1-0.3*ancho/largo

    return sc,sq,sg


def correccionTalud(beta,anguloRozamiento):

    if beta<=anguloRozamiento/2:

        # paso de grados a radianes
        betaRad=np.deg2rad(beta)
        anguloRozamientoRad=np.deg2rad(anguloRozamiento)

        # calculos de los coeficientes
        tc=np.exp(-2*betaRad*np.tan(anguloRozamientoRad))
        tq=1-np.sin(2*betaRad)
        tg=1-np.sin(2*betaRad)

    return tc,tq,tg

def densidad(pesoAparente,pesoSaturado,profApoyo,b,zw):
    
    # Cálculo de la densidad bajo cimentación en presencia del nivel freático

    if zw>=b:
        peso=pesoAparente
    if zw<=profApoyo:
        peso=pesoSaturado-9.81
    if profApoyo<zw<b:
        peso=(pesoSaturado-9.81)+(pesoAparente-(pesoSaturado-9.81))*zw/b

    return peso


